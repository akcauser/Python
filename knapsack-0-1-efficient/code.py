import openpyxl


# book isimli Workbook sınıfından türetilmiş nesneye istenilen .xclx uzantılı dosyanın ataması işlemi
book = openpyxl.load_workbook('Instance_24_6404180.xlsx')

# .xclx uzantılı book dosyasında kullanmayı istediğimiz sayfayı açma işlemi
worksheet = book['Sheet1']

# Veri sayısının veri seti üzerindeki sabit hücreden alınması
data_count_cell = worksheet['B1']
data_count = int(data_count_cell.value)
print("data count:", data_count)

# knapsack kapasitesinin veri seti üzerindeki sabit hücreden alınması
knapsack_capacity_cell = worksheet['B2']
knapsack_capacity = int(knapsack_capacity_cell.value)
print("knapsack capacity:", knapsack_capacity)

# veri depolaması için gerekli listelerin tanımlanması
data_profit = []
data_weight = []

# veri setindeki öngörülmüş hücrelerde bulunan 'profit' değişkeninin veri setiden alınma ve diziye ekleme işlemi
for row in worksheet.iter_rows(min_row=5, min_col=1, max_row=data_count + 4, max_col=1):
    for cell in row:
        # print(cell.value)#veri doğru diye kontrol ettik
        data_profit.append(int(cell.value))

# veri setindeki öngörülmüş hücrelerde bulunan 'weight' değişkeninin veri setiden alınma ve diziye ekleme işlemi
for row in worksheet.iter_rows(min_row=5, min_col=2, max_row=data_count + 4, max_col=2):
    for cell in row:
        # print(cell.value)#veri doğru mu diye kontrol ettik
        data_weight.append(int(cell.value))

# diziye eklediğimiz verilerin kontrolü ve kullanıcıya gösterilmesi
for i in range(0, (int(data_count))):
    print(data_profit[i], data_weight[i], sep="-")


# Yazdığımız algoritma kullanıcının girmiş olduğu veri setinde bulunan bütün maksimum verim olasılıklarını deneyerek
# arasından en verimli sonucu seçmesi işlemini yapmaktadır.
#Bu algoritmanın zaman karmaşıklığı (2^veri_sayısı)'dır.
def knapSack(capacity, weight, profit, count):
    if count == 0 or capacity == 0:
        return 0

    if (weight[count - 1] > capacity):
        return knapSack(capacity, weight, profit, count - 1)

    else:
        return max(profit[count - 1] + knapSack(capacity - weight[count - 1], weight, profit, count - 1)
                   , knapSack(capacity, weight, profit, count - 1))

#Yazdığımız algoritma veri seti için maximum sağlanabilecek verimi kullanıcıya döndürmektedir.
#Algoritmamızın zaman karmaşıklığı (veri sayısı * maksimum kapasite)'dir
def knapsackDP(capacity, weight, profit, count):

    mat = [[0 for i in range(capacity + 1)] for i in range(2)]

    item = 0
    while item < count:

        size = 0

        if item % 2 == 0:
            while size < capacity:
                size += 1
                if weight[item] <= size:
                    empty = size - weight[item]
                    mat[1][size] = max(profit[item] + mat[0][empty], mat[0][size])
                else:
                    mat[1][size] = mat[0][size]


        else:
            while size < capacity:
                size += 1
                if weight[item] <= size:
                    empty = size - weight[item]
                    mat[0][size] = max(profit[item] + mat[1][empty], mat[1][size])
                else:
                    mat[0][size] = mat[1][size]


        item += 1


    if count % 2 == 0:
        return mat[0][capacity]
    else:
        return mat[1][capacity]



# Optimal verimliliği hesaplayan knapsack isimli metodun gerekli parametreler ile çağırılması ve ekrana yazılması
print("optimal profit:")

#Burada yaptığımız kontrol zaman karmaşıklığı az olan knapsack çözümü hangisi ise onu kullanmamızı sağlıyor
if  pow(2,data_count) < (data_count*knapsack_capacity):
    print(knapSack(knapsack_capacity, data_weight, data_profit, data_count))
else:
    print(knapsackDP(knapsack_capacity, data_weight, data_profit, data_count))




