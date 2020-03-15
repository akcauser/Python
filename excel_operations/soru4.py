from openpyxl import Workbook

wb = Workbook()

wb.save('kitap_4.xlsx')

ws1  = wb.create_sheet("sayfa_1")
ws2  = wb.create_sheet("sayfa_2")
ws3  = wb.create_sheet("sayfa_4")

sheet = wb['Sheet']

wb.remove(sheet)

wb.save('kitap_4.xlsx')


