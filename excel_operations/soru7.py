from openpyxl import load_workbook

wb = load_workbook('kitap_6.xlsx')

ws = wb['sayfa_1']

for j in range(1,ws.max_column+1):
    for i in range(1,ws.max_row+1):
        if ws.cell(row=i, column=j).value > 100 and ws.cell(row=i, column=j).value < 1000:
            print(ws.cell(row=i, column=j).value)
