from openpyxl import Workbook

wb = Workbook()

ws1  = wb.create_sheet("sayfa_1")
ws2  = wb.create_sheet("sayfa_2")
ws3  = wb.create_sheet("sayfa_3")

first = ws3['A1']
first.value = '21312312 Sultan Akçor'

wb.save(filename = 'kitap_3.xlsx')


