from openpyxl import Workbook

wb = Workbook()

ws1  = wb.create_sheet("sayfa_1")
ws2  = wb.create_sheet("sayfa_2")

wb.save(filename = 'kitap_2.xlsx')


