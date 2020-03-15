from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Side, Border

wb = load_workbook('kitap_6.xlsx')

ws = wb['sayfa_1']
ws.title = 'tablo_1'

a10 = ws['A10']
a10.font = Font(color=colors.YELLOW, size=20)

b8 = ws['B8']
b8.font = Font(bold=True)

double = Side(border_style="double", color="ff0000")
c12 = ws['C12']
c12.border = Border(top=double, left=double, right=double, bottom=double)

wb.save('kitap_6.xlsx')