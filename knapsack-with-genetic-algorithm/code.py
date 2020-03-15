class Item(object):
    def __init__(self, v, w):
        self.value = v
        self.weight = w

import random
import openpyxl

#test için dosya isminin değiştirilmesi yeterlidir.
# book isimli Workbook sınıfından türetilmiş nesneye istenilen .xclx uzantılı dosyanın ataması işlemi
book = openpyxl.load_workbook('Instance_3_50.xlsx')

# .xclx uzantılı book dosyasında kullanmayı istediğimiz sayfayı açma işlemi
worksheet = book['Sheet1']

# Veri sayısının veri seti üzerindeki sabit hücreden alınması
data_count_cell = worksheet['B1']
data_count = int(data_count_cell.value)
print("data count:", data_count)

# knapsack kapasitesinin veri seti üzerindeki sabit hücreden alınması
knapsack_capacity_cell = worksheet['B2']
CAPACITY = int(knapsack_capacity_cell.value)
print("knapsack capacity:", CAPACITY)

# veri depolaması için gerekli listelerin tanımlanması
dataset_profit = []
dataset_weight = []
dataset_perprofit = []

# veri setindeki öngörülmüş hücrelerde bulunan 'profit' değişkeninin veri setiden alınma ve diziye ekleme işlemi
for row in worksheet.iter_rows(min_row=5, min_col=1, max_row=data_count + 4, max_col=1):
    for cell in row:
        # print(cell.value)#veri doğru diye kontrol ettik
        dataset_profit.append(int(cell.value))

# veri setindeki öngörülmüş hücrelerde bulunan 'weight' değişkeninin veri setiden alınma ve diziye ekleme işlemi
for row in worksheet.iter_rows(min_row=5, min_col=2, max_row=data_count + 4, max_col=2):
    for cell in row:
        # print(cell.value)#veri doğru mu diye kontrol ettik
        dataset_weight.append(int(cell.value))

for index in range(data_count):
    dataset_perprofit.append(dataset_profit[index] / dataset_weight[index])

# veriler excel dosyasından çekilerek ITEMS dizisine eklenecek
ITEMS = [Item(dataset_profit[x], dataset_weight[x]) for x in range(0, data_count)]

#populasyon boyutu - populasyondaki kromozom sayısı
POP_SIZE = 100

# Üzerinde çalışmasını istediğimiz gen sayısı
GEN_MAX = 500

#verimlilik hesaplama fonksiyonu
def fitness(target):
    total_value = 0
    total_weight = 0
    index = 0
    for i in target:
        if index >= len(ITEMS):
            break
        if (i == 1):
            total_value += ITEMS[index].value
            total_weight += ITEMS[index].weight
        index += 1

    if total_weight > CAPACITY:
        # kapasiteyi aşıyorsa
        return 0
    else:
        # kapasiteyi aşmıyorsa
        return total_value

#ilk popolasyonun oluşturulması fonksiyonu
def create_population(amount):
    generation1 = []
    gen = []
    for x in range(amount):
        if x == 0:
            generation1.append(create_chromosome_greedy())
        else:
            gen = create_chromosome()

        if fitness(gen) == 0:
            generation1.append(create_chromosome_greedy())
        else:
            generation1.append(gen)

    return generation1

#ilk kromozomun oluşturulması, greedy approach yöntemiyle
def create_chromosome_greedy():
    chromosome = [0 for _ in range(data_count)]

    doluluk = 0
    for _ in dataset_profit:

        if (CAPACITY - doluluk) < dataset_weight[dataset_perprofit.index(max(dataset_perprofit))]:
            continue

        doluluk += dataset_weight[dataset_perprofit.index(max(dataset_perprofit))]
        chromosome[dataset_perprofit.index(max(dataset_perprofit))] = 1
        dataset_perprofit[dataset_perprofit.index(max(dataset_perprofit))] = 0

    return chromosome

#ilk kromozomun oluşturulması random
def create_chromosome():
    return [random.randint(0, 1) for _ in range(0, len(ITEMS))]

#mutasyon fonksiyonu
def mutate(target):
    r = random.randint(0, len(target) - 1)
    if target[r] == 1:
        target[r] = 0
    else:
        target[r] = 1

#1.nesil evrimleştirilerek 2. nesile ulaşılması
def evolve_population(pop):

    parent_eligibility = 0.1
    mutation_chance = random.random()
    parent_lottery = 0.05

    #parent_length = round(parent_eligibility * data_count) + 1
    parent_length = int(parent_eligibility * len(pop))
    parent_length = 10

    parents = pop[:parent_length]
    nonparents = pop[parent_length:]
    #parent olamayacaklar arasından rastgele kura seçimi ile parent olabileceklere ekleme yaptık
    for np in nonparents:
        if parent_lottery > random.random():
            parents.append(np)

    #parent olabilecekler arasından random seçim ile mutasyon gerçekleştirdik
    for p in parents:
        if mutation_chance > random.random():
            mutate(p)


    children = []
    desired_length = len(pop) - len(parents)
    while len(children) < desired_length:

        male = parents[random.randint(0, len(parents) - 1)]
        female = parents[random.randint(0, len(parents) - 1)]
        half = int(len(male) / 2)

        child = male[:half] + female[half:]
        if mutation_chance > random.random():
            mutate(child)
        children.append(child)


    parents.extend(children)

    return parents

##MAIN kısmı. Fonkiyonların çağırılması
generation = 1
population = create_population(POP_SIZE)
best = 0

# last = list()

for g in range(0, GEN_MAX):
    print("Generation ", generation, " with ", len(population))
    population = sorted(population, key=lambda x: fitness(x), reverse=True)
    for i in population:
        if fitness(i) != 0:
            print(str(i), ", fit:", fitness(i))
            if fitness(i) > best:
                # last.clear()
                # for bit in i:
                #     last.append(bit)
                best = fitness(i)

        else:
            print(str(i), ", fit: Capacity Over")

    population = evolve_population(population)

    generation += 1

print("approximate value to optimal profit : ", best)


#print(last)